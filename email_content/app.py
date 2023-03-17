import boto3
from datetime import date
from openpyxl import load_workbook
import io


def send_emails(event, context):

    # Set up SES client
    ses_client = boto3.client('ses', region_name='us-east-1')

    # Setup s3 Client
    s3_client = boto3.client('s3')

    # Get s3 bucket and object key from the event object
    bucket_name = event['Records'][0]['s3']['bucket']['name']
    object_key = event['Records'][0]['s3']['object']['key']

    # Read Excel file from S3
    response = s3_client.get_object(Bucket=bucket_name, Key=object_key)
    file_content = response['Body'].read()

    # Load workbook and get the first worksheet
    workbook = load_workbook(io.BytesIO(file_content), read_only=True)
    sheet = workbook.active

    # Iterate through the rows of the worksheet and send an email to each contact
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if Contact Information is not blank
        contact_information = row[10]  # Column K

        if contact_information:
            # Get email and URL from the row
            email = contact_information
            url = row[3]  # Column D
            subject = row[0] if row[0] else "Default Subject"  # Column A

            # Construct email message
            subject = subject
            body_text = f'Hello,\n\nPlease visit this URL: {url}\n\nThank you!'
            sender = 'jklacyn@amazon.com'
            recipient = email

            # Send email using SES client
            response = ses_client.send_email(
                Destination={
                    'ToAddresses': [
                        recipient,
                    ],
                },
                Message={
                    'Body': {
                        'Text': {
                            'Charset': 'UTF-8',
                            'Data': body_text,
                        },
                    },
                    'Subject': {
                        'Charset': 'UTF-8',
                        'Data': subject,
                    },
                },
                Source=sender,
            )

            # Print response from SES client
            print(f"Email sent to {email} with response: {response}")
